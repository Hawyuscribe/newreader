"""
Export vascular neurology MCQs to Excel format with multiple sheets for better analysis.
"""
import json
from django.core.management.base import BaseCommand
from mcq.models import MCQ
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class Command(BaseCommand):
    help = 'Export vascular neurology/stroke MCQs to Excel format with detailed analysis'

    def add_arguments(self, parser):
        parser.add_argument(
            '--filename',
            type=str,
            default=f'vascular_mcqs_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            help='Output filename for the Excel export'
        )

    def handle(self, *args, **options):
        filename = options['filename']
        
        # Query for vascular neurology MCQs
        vascular_mcqs = MCQ.objects.filter(
            subspecialty__in=['Vascular Neurology/Stroke', 'Vascular Neurology', 'Stroke']
        ).order_by('exam_type', 'exam_year', 'question_number')
        
        self.stdout.write(f"Found {vascular_mcqs.count()} vascular neurology MCQs")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        self.create_summary_sheet(wb, vascular_mcqs)
        
        # Create detailed MCQs sheet
        self.create_detailed_sheet(wb, vascular_mcqs)
        
        # Create sheets by exam type and year
        exam_groups = {}
        for mcq in vascular_mcqs:
            key = f"{mcq.exam_type or 'Unknown'}_{mcq.exam_year or 'Unknown'}"
            if key not in exam_groups:
                exam_groups[key] = []
            exam_groups[key].append(mcq)
        
        for exam_key, mcqs in sorted(exam_groups.items()):
            self.create_exam_sheet(wb, exam_key, mcqs)
        
        # Save workbook
        wb.save(filename)
        self.stdout.write(self.style.SUCCESS(f"Successfully exported to {filename}"))
    
    def create_summary_sheet(self, wb, mcqs):
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary")
        
        # Headers
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Summary data
        summary_data = [
            ["Total MCQs", len(mcqs)],
            ["MCQs with Explanations", sum(1 for mcq in mcqs if mcq.explanation_sections)],
            ["MCQs with Images", sum(1 for mcq in mcqs if mcq.image_url)],
        ]
        
        # Add exam type breakdowns
        exam_counts = {}
        for mcq in mcqs:
            key = f"{mcq.exam_type or 'Unknown'} - {mcq.exam_year or 'Unknown'}"
            exam_counts[key] = exam_counts.get(key, 0) + 1
        
        summary_data.append(["", ""])  # Empty row
        summary_data.append(["Breakdown by Exam", "Count"])
        
        for exam_key, count in sorted(exam_counts.items()):
            summary_data.append([exam_key, count])
        
        # Write data
        for row_idx, row_data in enumerate(summary_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_detailed_sheet(self, wb, mcqs):
        """Create detailed MCQs sheet with all data"""
        ws = wb.create_sheet("All MCQs")
        
        # Headers
        headers = [
            'ID', 'Question Number', 'Exam Type', 'Exam Year', 'Question Text',
            'Option A', 'Option B', 'Option C', 'Option D', 'Option E', 'Option F',
            'Correct Answer', 'Subspecialty', 'Has Image', 'Image URL',
            'Conceptual Foundation', 'Pathophysiology', 'Clinical Correlation',
            'Diagnostic Approach', 'Classification', 'Management Principles',
            'Option Analysis', 'Clinical Pearls', 'Current Evidence'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write MCQ data
        for row_idx, mcq in enumerate(mcqs, 2):
            options = mcq.get_options_dict()
            explanation_sections = mcq.explanation_sections or {}
            
            # Helper function to get section content
            def get_section(keys):
                for key in keys:
                    if key in explanation_sections and explanation_sections[key]:
                        return explanation_sections[key]
                return ''
            
            row_data = [
                mcq.id,
                mcq.question_number or '',
                mcq.exam_type or '',
                mcq.exam_year or '',
                mcq.question_text,
                options.get('A', ''),
                options.get('B', ''),
                options.get('C', ''),
                options.get('D', ''),
                options.get('E', ''),
                options.get('F', ''),
                mcq.correct_answer,
                mcq.subspecialty,
                'Yes' if mcq.image_url else 'No',
                mcq.image_url or '',
                get_section(['conceptual foundation', 'conceptual_foundation']),
                get_section(['pathophysiology', 'pathophysiological_mechanisms']),
                get_section(['clinical correlation', 'clinical_correlation', 'clinical context']),
                get_section(['diagnostic approach', 'diagnostic_approach']),
                get_section(['classification and neurology', 'classification_neurology']),
                get_section(['management principles', 'management_principles']),
                get_section(['option analysis', 'option_analysis']),
                get_section(['clinical pearls', 'clinical_pearls', 'key insight']),
                get_section(['current evidence', 'current_evidence', 'quick reference'])
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Wrap text for long content
                if col_idx in [5, 16, 17, 18, 19, 20, 21, 22, 23, 24]:  # Text-heavy columns
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Adjust column widths
        self.adjust_column_widths(ws)
    
    def create_exam_sheet(self, wb, exam_key, mcqs):
        """Create sheet for specific exam type and year"""
        # Clean sheet name (Excel has restrictions)
        sheet_name = exam_key.replace('/', '_')[:31]  # Excel sheet names max 31 chars
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = [
            'Question Number', 'Question Text', 'Options', 'Correct Answer',
            'Has Explanation', 'Has Image'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Write MCQ data
        for row_idx, mcq in enumerate(mcqs, 2):
            options = mcq.get_options_dict()
            options_text = '\n'.join([f"{k}: {v}" for k, v in sorted(options.items())])
            
            row_data = [
                mcq.question_number or f"ID:{mcq.id}",
                mcq.question_text,
                options_text,
                mcq.correct_answer,
                'Yes' if mcq.explanation_sections else 'No',
                'Yes' if mcq.image_url else 'No'
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx in [2, 3]:  # Text-heavy columns
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        self.adjust_column_widths(ws)
    
    def adjust_column_widths(self, ws):
        """Adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set reasonable limits
            if column[0].column in [5, 16, 17, 18, 19, 20, 21, 22, 23, 24]:  # Text columns
                adjusted_width = min(max_length * 0.3, 60)
            else:
                adjusted_width = min(max_length + 2, 30)
            
            ws.column_dimensions[column_letter].width = adjusted_width